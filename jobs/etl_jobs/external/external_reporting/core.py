import warnings
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional

import duckdb
import openpyxl
from google.cloud import bigquery
from treelib import Tree

from config import (
    BIGQUERY_ANALYTICS_DATASET,
    GCP_PROJECT,
    REPORTS,
    SHEET_DEFINITIONS,
    SOURCE_TABLES,
    STAKEHOLDER_REPORTS,
)
from services.tracking import ReportStats
from utils.data_utils import (
    build_region_hierarchy,
    sanitize_date_fields,
    sanitize_numeric_types,
)
from utils.verbose_logger import log_print

# log_print = logging.getLogger(__name__)
# logging.basicConfig(level=logging.INFO)


class StakeholderType(Enum):
    MINISTERE = "ministere"
    DRAC = "drac"


@dataclass
class Stakeholder:
    """Stakeholder - either Ministère (no children) or DRAC (region → academies & departments)."""

    type: StakeholderType
    name: str
    desired_reports: List[str] = field(default_factory=list)
    academy_tree: Optional[Tree] = None
    department_tree: Optional[Tree] = None
    hierarchy_data: Optional[Dict] = field(default=None, repr=False)
    build_trees: bool = field(default=True)

    def __post_init__(self):
        # Assign default reports
        if not self.desired_reports:
            self.desired_reports = list(STAKEHOLDER_REPORTS.get(self.type.value, []))

        # Only DRAC stakeholders get trees, and only if requested
        if self.type == StakeholderType.DRAC and self.build_trees:
            # Fetch hierarchy from BigQuery or use provided one
            if self.hierarchy_data:
                hierarchy = self.hierarchy_data
            else:
                hierarchy = build_region_hierarchy()

            # Ensure region exists
            if self.name not in hierarchy:
                raise ValueError(f"❌ Region '{self.name}' not found in hierarchy")

            scope = hierarchy[self.name]

            # Build departement tree
            self.department_tree = Tree()
            self.department_tree.create_node("Departements", "root_dept")
            for dep in scope["departements"]:
                self.department_tree.create_node(
                    dep, f"{self.name}_dep_{dep}", parent="root_dept"
                )

            # Build academy tree
            self.academy_tree = Tree()
            self.academy_tree.create_node("Academies", "root_acad")
            for acad in scope["academies"]:
                self.academy_tree.create_node(
                    acad, f"{self.name}_acad_{acad}", parent="root_acad"
                )

        else:
            self.academy_tree = None
            self.department_tree = None


@dataclass
class Sheet:
    """Represents a sheet instance within a report."""

    definition: str  # key from SHEET_DEFINITIONS
    template_tab: str  # template tab name
    worksheet: Any  # openpyxl worksheet instance
    filters: Dict[str, Any] = field(default_factory=dict)
    tab_name: Optional[str] = None
    stakeholder: Optional[Any] = None  # reference to Stakeholder
    context: Optional[Dict[str, Any]] = None  # region, academy, department, etc.
    layout: Optional[Dict[str, Any]] = None

    def get_dimension_context(self) -> Optional[Dict[str, str]]:
        """
        Resolve dimension context for DuckDB queries based on filters and context.

        Returns:
            Dict with 'name' (dimension name) and 'value' (dimension value) for DuckDB queries
        """
        try:
            scale = self.filters.get("scale", "national")

            # Map scale to dimension name
            dimension_name_map = {
                "national": "NAT",
                "region": "REG",
                "academie": "ACAD",
                "departement": "DEP",
            }

            dimension_name = dimension_name_map.get(scale, "NAT")

            # Get dimension value from context
            if scale == "national":
                dimension_value = "NAT"
            else:
                dimension_value = self.context.get(scale) if self.context else None

            if not dimension_value and scale != "national":
                log_print.warning(
                    f"Could not resolve dimension value for scale '{scale}' in context {self.context}"
                )
                return None

            return {"name": dimension_name, "value": dimension_value}

        except Exception as e:
            log_print.warning(f"Failed to resolve dimension context: {e}")
            return None


class Report:
    """Represents a single Excel report for a stakeholder."""

    def __init__(
        self,
        report_type: str,
        stakeholder,
        base_template_path: Path,
        output_path: Path,
        context: Dict[str, Any] = None,
    ):
        self.report_type = report_type
        self.stakeholder = stakeholder
        self.base_template_path = base_template_path
        self.output_path = output_path
        self.context = context or {}
        self.workbook: Any = None  # openpyxl.Workbook
        self.sheets: List[Sheet] = []  # list of Sheet instances

    def _load_template(self):
        """Load the template workbook directly as the report workbook."""
        self.workbook = openpyxl.load_workbook(self.base_template_path)

        # Remove default sheet if it exists and is empty
        if (
            "Sheet" in self.workbook.sheetnames
            and len(self.workbook["Sheet"].rows) == 0
        ):
            self.workbook.remove(self.workbook["Sheet"])

    def _build_tab_name(
        self, definition_key: str, filters: Dict[str, Any], context: Dict[str, Any]
    ):
        """Return a user-friendly tab name based on sheet definition, filters, and context."""

        if definition_key == "individual_kpis":
            base = "Part Individuelle"
            scale = filters.get("scale", "").capitalize()
            return f"{base} ({scale})" if scale else base

        elif definition_key == "collective_kpis":
            base = "Part Collective"
            scale = filters.get("scale", "").capitalize()
            return f"{base} ({scale})" if scale else base

        else:
            tab_name = SHEET_DEFINITIONS.get(definition_key, {}).get("tab_name")
            if tab_name is None:
                tab_name = SHEET_DEFINITIONS.get(definition_key, {}).get(
                    "template_tab", definition_key
                )
            return tab_name.capitalize()

    def _resolve_sheets(self):
        """Create Sheet objects from template tabs based on report blueprint."""
        blueprint = REPORTS[self.report_type]

        for sheet_info in blueprint["sheets"]:
            definition_key = sheet_info["definition"]
            template_info = SHEET_DEFINITIONS[definition_key]

            # Decide tab name based on definition + filters + context
            tab_name = self._build_tab_name(
                definition_key,
                {**sheet_info.get("filters", {}), **self.context},
                self.context,
            )

            # Check if the tab already exists
            if tab_name in self.workbook.sheetnames:
                ws = self.workbook[tab_name]
            else:
                # Copy the template tab
                template_ws = self.workbook[template_info["template_tab"]]
                ws = self.workbook.copy_worksheet(template_ws)

            if len(tab_name) > 31:
                log_print.warning(
                    f"⚠️ Tab name '{tab_name}' exceeds 31 characters, truncating.",
                    fg="yellow",
                )
                tab_name = tab_name[:31]  # truncate if too long
            ws.title = tab_name

            # Create Sheet instance
            sheet = Sheet(
                definition=definition_key,
                template_tab=template_info["template_tab"],
                worksheet=ws,
                filters={**sheet_info.get("filters", {}), **self.context},
                tab_name=tab_name,
                stakeholder=self.stakeholder,
                context=self.context,
            )
            self.sheets.append(sheet)

    def _cleanup_template_sheets(self, used_tabs: List[str]):
        """Remove all template/unused sheets and enforce blueprint order."""
        # Remove any sheet not in the used list
        for sheet_name in self.workbook.sheetnames:
            if sheet_name not in used_tabs:
                self.workbook.remove(self.workbook[sheet_name])

        # Enforce order from used_tabs
        for idx, tab_name in enumerate(used_tabs):
            ws = self.workbook[tab_name]
            self.workbook._sheets.remove(ws)
            self.workbook._sheets.insert(idx, ws)

    def build(self, ds: str, duckdb_conn: duckdb.DuckDBPyConnection) -> ReportStats:
        """
        Build the report workbook using the new service architecture.

        Args:
            ds: Consolidation date in YYYY-MM-DD format
            duckdb_conn: DuckDB connection for data queries

        Returns:
            ReportStats with detailed processing statistics
        """
        from services.orchestration import ReportOrchestrationService

        log_print.info(
            f"⚙️  Building report: {self.report_type} for {self.stakeholder.name}",
            fg="cyan",
        )

        # Create report stats object
        report_stats = ReportStats(
            report_name=self.output_path.name, report_type=self.report_type
        )

        # Step 1: Load template and create sheets
        self._load_template()
        self._resolve_sheets()
        self._cleanup_template_sheets([sheet.tab_name for sheet in self.sheets])

        # Step 2: Use ReportOrchestrationService for complex processing
        orchestrator = ReportOrchestrationService(duckdb_conn)
        sheet_stats_list = orchestrator.process_all_sheets(self.sheets, ds)

        # Add all sheet stats to report stats
        for sheet_stats in sheet_stats_list:
            report_stats.add_sheet_stats(sheet_stats)

        # Step 3: Log final statistics
        if (
            report_stats.kpis_successful == report_stats.total_kpis
            and report_stats.tops_successful == report_stats.total_tops
        ):
            log_print.info(
                f"✅ Report completed successfully: {self.output_path.name}", fg="green"
            )
        elif report_stats.kpis_successful > 0 or report_stats.tops_successful > 0:
            log_print.warning(
                f"⚠️ Report completed with warnings: "
                f"KPIs: {report_stats.kpis_successful}/{report_stats.total_kpis}, "
                f"Tops: {report_stats.tops_successful}/{report_stats.total_tops}",
                fg="yellow",
            )
        else:
            log_print.error(
                "❌ Report failed: No data processed successfully", fg="red"
            )

        return report_stats

    def save(self):
        """Save the workbook to the output path."""
        if self.output_path.suffix != ".xlsx":
            self.output_path = self.output_path.with_suffix(".xlsx")
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(self.output_path)
        log_print.info(f"💾 Saved report: {self.output_path}")


class ReportPlanner:
    """Expands a stakeholder into the list of reports they should get."""

    def __init__(self, stakeholder):
        self.stakeholder = stakeholder

    def plan_reports(self) -> List[Dict[str, Any]]:
        """
        Returns a list of report jobs:
        {
            "report_type": str,
            "context": dict,  # region, academy, department, etc.
            "output_path": Path
        }
        """
        jobs = []

        if self.stakeholder.type.name == "MINISTERE":
            # Ministère gets only the national summary
            # base_path = Path("NATIONAL")
            jobs.append(
                {
                    "report_type": "national_summary",
                    "context": {},
                    "output_path": "rapport_national.xlsx",
                }
            )
            return jobs

        # DRAC gets:
        # 1) Regional summary (includes national for comparison)
        region_name = self.stakeholder.name
        jobs.append(
            {
                "report_type": "region_summary",
                "context": {"region": region_name},
                "output_path": "rapport_regional.xlsx",
            }
        )

        # 2) One detailed report per académie (collective scope)
        if self.stakeholder.academy_tree:
            for node in self.stakeholder.academy_tree.all_nodes():
                if node.tag not in ["root_acad", "Academies"]:
                    jobs.append(
                        {
                            "report_type": "academy_detail",
                            "context": {"region": region_name, "academie": node.tag},
                            "output_path": f"academie_{node.tag}.xlsx",
                        }
                    )

        # 3) One detailed report per département (individual scope)
        if self.stakeholder.department_tree:
            for node in self.stakeholder.department_tree.all_nodes():
                if node.tag not in ["root_dept", "Departements"]:
                    jobs.append(
                        {
                            "report_type": "department_detail",
                            "context": {"region": region_name, "departement": node.tag},
                            "output_path": f"departement_{node.tag}.xlsx",
                        }
                    )

        return jobs


def process_report_worker(task: Dict[str, Any]) -> ReportStats:
    """
    Worker function to process a single report.
    """
    try:
        # Reconstruct Stakeholder (without trees for memory efficiency)
        stakeholder = Stakeholder(
            type=StakeholderType(task["stakeholder_type"]),
            name=task["stakeholder_name"],
            build_trees=False,
        )

        # Connect to DB (ReadOnly)
        conn = duckdb.connect(database=str(task["db_path"]), read_only=True)

        try:
            # Create Report instance
            # Note: The template should have been copied to output_path already by main process
            report = Report(
                report_type=task["report_type"],
                stakeholder=stakeholder,
                base_template_path=Path(task["output_path"]),
                output_path=Path(task["output_path"]),
                context=task["context"],
            )

            # Build
            report_stats = report.build(task["ds"], conn)

            # Save (overwrite self)
            report.save()

            log_print.debug(f"🏁 Worker finished logic for {task.get('output_path')}")
            return report_stats

        finally:
            log_print.debug(f"🔌 Closing DB connection for {task.get('output_path')}")
            conn.close()
            log_print.debug(f"👋 Worker exiting for {task.get('output_path')}")

    except Exception as e:
        log_print.error(f"❌ Worker failed for {task.get('output_path')}: {e}")
        # Return failed stats
        return ReportStats(
            report_name=str(task.get("output_path", "unknown")),
            report_type=task.get("report_type", "unknown"),
        )


class ExportSession:
    """Manages the entire export session with proper resource cleanup."""

    def __init__(self, ds: str, db_path: str = ":memory:"):
        self.ds = ds
        self.tables_to_load = SOURCE_TABLES
        self.db_path = db_path
        self.conn: Optional[duckdb.DuckDBPyConnection] = None

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.conn:
            self.conn.close()

    def load_data(self):
        """Load tables based on self.scope."""

        client = bigquery.Client(project=GCP_PROJECT)
        # Connect to persistent file if specified
        self.conn = duckdb.connect(self.db_path)

        for _, config in self.tables_to_load.items():
            self._load_table(client, config)

    def _load_table(self, client: bigquery.Client, config: Dict):
        """Load a single table into DuckDB."""
        table_name = config["table"]
        dataset = config.get("dataset", BIGQUERY_ANALYTICS_DATASET)
        project_id = config.get("project", GCP_PROJECT)
        log_print.info(f"⏳ Loading {table_name} from BigQuery...")
        query = f"SELECT * FROM `{project_id}.{dataset}.{table_name}`"
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore", message=".*BigQuery Storage module not found.*"
            )
            df = client.query(query).to_dataframe()
            log_print.debug(f"{table_name} dtypes: {df.dtypes}")
        if df.empty:
            log_print.warning(f"Warning: {table_name} returned no data")
            return

        df = sanitize_date_fields(df, "partition_month")
        df = sanitize_numeric_types(df)

        log_print.info(f"✅ Loaded {len(df):,} records for {table_name}")

        # Register and create table
        self.conn.register(f"{table_name}_tmp", df)
        self.conn.execute(
            f"CREATE TABLE {table_name} AS SELECT * FROM {table_name}_tmp"
        )

        # Cleanup DataFrame immediately
        self.conn.unregister(f"{table_name}_tmp")
        del df
        import gc

        gc.collect()

        # Create indexes based on available columns
        self._create_indexes(table_name)
        log_print.info(f"✅ Created {table_name} table")

    def _create_indexes(self, table_name: str):
        """Create indexes on commonly queried columns with error recovery."""
        # Fetch columns from DuckDB
        columns = [
            row[0] for row in self.conn.execute(f"DESCRIBE {table_name}").fetchall()
        ]

        indexes_created = []
        indexes_failed = []

        log_print.info(f"🔍 Creating indexes for table: {table_name}")

        # List of indexes to attempt
        index_operations = []

        # Core filtering columns
        core_filter_columns = ["dimension_name", "dimension_value", "partition_month"]
        kpi_columns = ["kpi_name"]

        if all(col in columns for col in core_filter_columns):
            index_operations.append(
                {
                    "name": "core_filters",
                    "sql": f"CREATE INDEX IF NOT EXISTS idx_{table_name}_core_filters ON {table_name} (dimension_name, dimension_value, partition_month)",
                    "description": "Core filters composite index",
                }
            )

        if all(col in columns for col in core_filter_columns + kpi_columns):
            index_operations.append(
                {
                    "name": "kpi_filters",
                    "sql": f"CREATE INDEX IF NOT EXISTS idx_{table_name}_kpi_filters ON {table_name} (kpi_name, dimension_name, dimension_value, partition_month)",
                    "description": "KPI-specific composite index",
                }
            )

        # Individual indexes
        for col in ["partition_month", "dimension_name", "dimension_value", "kpi_name"]:
            if col in columns:
                index_operations.append(
                    {
                        "name": col,
                        "sql": f"CREATE INDEX IF NOT EXISTS idx_{table_name}_{col} ON {table_name} ({col})",
                        "description": f"Individual index on {col}",
                    }
                )

        # Execute each index creation with individual error handling
        for idx_op in index_operations:
            try:
                self.conn.execute(idx_op["sql"])
                indexes_created.append(idx_op["name"])
                log_print.debug(f"   ✓ {idx_op['description']}")
            except Exception as e:
                indexes_failed.append(idx_op["name"])
                if "INTERNAL Error" in str(e):
                    log_print.warning(
                        f"   ⚠ {idx_op['description']} - skipped (DuckDB internal error)",
                        fg="yellow",
                    )
                else:
                    log_print.error(
                        f"   ❌ {idx_op['description']} - failed: {str(e)[:50]}...",
                        fg="red",
                    )

        # Summary
        if indexes_created:
            log_print.debug(
                f"📊 {table_name}: Created {len(indexes_created)} indexes successfully",
                fg="green",
            )
        if indexes_failed:
            log_print.warning(
                f"⚠ {table_name}: {len(indexes_failed)} indexes failed (table will still work)",
                fg="yellow",
            )
