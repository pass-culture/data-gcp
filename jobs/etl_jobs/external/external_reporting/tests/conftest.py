"""Fixtures and test data for the external_reporting test suite."""

from datetime import date
from unittest.mock import Mock, patch

import pandas as pd
import pytest
from openpyxl import Workbook


@pytest.fixture(scope="session", autouse=True)
def mock_gcp_services():
    """
    Global mock for GCP services to avoid authentication issues during tests.
    This fixture automatically applies to all tests.
    """
    with (
        patch("google.cloud.bigquery.Client") as mock_bq,
        patch("google.cloud.storage.Client") as mock_storage,
    ):
        # Configure mocks to return appropriate test values
        mock_bq_instance = Mock()
        mock_bq.return_value = mock_bq_instance

        mock_storage_instance = Mock()
        mock_storage.return_value = mock_storage_instance

        yield {
            "bigquery": mock_bq_instance,
            "storage": mock_storage_instance,
        }


@pytest.fixture()
def mock_bigquery_client():
    """Mock BigQuery client for testing."""
    mock_client = Mock()
    mock_client.project = "test-project"

    # Mock query results
    mock_query_job = Mock()
    mock_query_job.result.return_value = iter([])
    mock_client.query.return_value = mock_query_job

    return mock_client


@pytest.fixture()
def mock_duckdb_connection():
    """Mock DuckDB connection for testing."""
    mock_conn = Mock()

    # Mock execute and cursor methods to return a mock result with df() method
    mock_result = Mock()
    mock_result.df.return_value = pd.DataFrame()
    mock_conn.execute.return_value = mock_result
    mock_conn.cursor.return_value = mock_conn

    return mock_conn


@pytest.fixture()
def sample_kpi_data():
    """Sample KPI data for testing."""
    return pd.DataFrame(
        {
            "partition_month": pd.to_datetime(
                ["2021-01-01", "2021-02-01", "2021-03-01", "2021-12-01"]
            ),
            "kpi_name": ["test_kpi"] * 4,
            "dimension_name": ["NAT"] * 4,
            "dimension_value": ["NAT"] * 4,
            "numerator": [100.0, 150.0, 200.0, 500.0],
            "denominator": [1000.0, 1000.0, 1000.0, 1000.0],
            "kpi": [0.1, 0.15, 0.2, 0.5],
        }
    )


@pytest.fixture()
def sample_kpi_data_with_year_label():
    """Sample KPI data with year_label for testing yearly aggregations."""
    return pd.DataFrame(
        {
            "partition_month": pd.to_datetime(
                [
                    "2021-01-01",
                    "2021-06-01",
                    "2021-12-01",
                    "2022-01-01",
                    "2022-06-01",
                    "2022-12-01",
                ]
            ),
            "kpi_name": ["test_kpi"] * 6,
            "year_label": [2021, 2021, 2021, 2022, 2022, 2022],
            "numerator": [100.0, 150.0, 200.0, 300.0, 350.0, 400.0],
            "denominator": [1000.0, 1000.0, 1000.0, 1000.0, 1000.0, 1000.0],
            "kpi": [0.1, 0.15, 0.2, 0.3, 0.35, 0.4],
        }
    )


@pytest.fixture()
def sample_top_data():
    """Sample top rankings data for testing."""
    return pd.DataFrame(
        {
            "partition_month": [
                pd.Timestamp("2024-10-01"),
                pd.Timestamp("2024-10-01"),
                pd.Timestamp("2024-10-01"),
            ],
            "offer_name": ["Offer A", "Offer B", "Offer C"],
            "bookings_count": [100, 75, 50],
            "rank": [1, 2, 3],
        }
    )


@pytest.fixture()
def sample_hierarchy_data():
    """Sample region hierarchy data for testing."""
    return {
        "Île-de-France": {
            "departements": ["Paris", "Hauts-de-Seine", "Seine-Saint-Denis"],
            "academies": ["Paris", "Créteil", "Versailles"],
            "academy_departments": {
                "Paris": ["Paris"],
                "Créteil": ["Seine-Saint-Denis"],
                "Versailles": ["Hauts-de-Seine"],
            },
        },
        "Auvergne-Rhône-Alpes": {
            "departements": ["Rhône", "Isère", "Loire"],
            "academies": ["Lyon", "Grenoble"],
            "academy_departments": {"Lyon": ["Rhône", "Loire"], "Grenoble": ["Isère"]},
        },
    }


@pytest.fixture()
def sample_bigquery_hierarchy_rows():
    """Sample BigQuery rows for hierarchy testing."""

    class MockRow:
        def __init__(self, region_name, dep_name, academy_name):
            self.region_name = region_name
            self.dep_name = dep_name
            self.academy_name = academy_name

    return [
        MockRow("Île-de-France", "Paris", "Paris"),
        MockRow("Île-de-France", "Hauts-de-Seine", "Versailles"),
        MockRow("Île-de-France", "Seine-Saint-Denis", "Créteil"),
        MockRow("Auvergne-Rhône-Alpes", "Rhône", "Lyon"),
        MockRow("Auvergne-Rhône-Alpes", "Loire", "Lyon"),
        MockRow("Auvergne-Rhône-Alpes", "Isère", "Grenoble"),
    ]


@pytest.fixture()
def mock_openpyxl_workbook():
    """Mock openpyxl workbook for testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Sheet"

    # Add some test data structure
    ws["A1"] = "kpi_name=test_kpi"
    ws["B1"] = "kpi"
    ws["C1"] = "sum"

    return wb


@pytest.fixture()
def mock_openpyxl_worksheet(mock_openpyxl_workbook):
    """Mock openpyxl worksheet for testing."""
    return mock_openpyxl_workbook.active


@pytest.fixture()
def sample_date_mappings():
    """Sample date mappings for Excel column expansion."""
    return {
        "years": [
            {1: "2021"},
            {2: "2022"},
            {3: "2023"},
        ],
        "months": [
            {4: "10/2024"},
            {5: "11/2024"},
            {6: "12/2024"},
        ],
    }


@pytest.fixture()
def sample_kpi_data_dict():
    """Sample KPI data dictionary for Excel writing."""
    return {
        "yearly": {
            2021: 100.5,
            2022: 150.3,
            2023: 200.7,
        },
        "monthly": {
            "10/2024": 15.2,
            "11/2024": 18.5,
            "12/2024": 20.1,
        },
    }


@pytest.fixture()
def sample_ds():
    """Sample consolidation date for testing."""
    return "2024-10-01"


@pytest.fixture()
def sample_ds_date():
    """Sample consolidation date as date object."""
    return date(2024, 10, 1)
