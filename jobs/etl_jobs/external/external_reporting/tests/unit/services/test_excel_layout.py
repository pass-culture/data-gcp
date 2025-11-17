"""Tests for ExcelLayoutService class."""

from unittest.mock import patch

from openpyxl import Workbook

from services.excel_layout import ExcelLayoutService


class TestExpandDateColumnsKpis:
    """Test cases for expand_date_columns_kpis method."""

    def test_expand_date_columns_indiv_delegates_correctly(self):
        """Test individual KPI column expansion delegates to internal method."""
        # Arrange
        wb = Workbook()
        ws = wb.active

        # Act - Mock the internal method to avoid complex setup
        with patch(
            "services.excel_layout.ExcelLayoutService._expand_date_columns_indiv"
        ) as mock_expand:
            mock_expand.return_value = {
                "date_mappings": {"years": [], "months": []},
                "total_width": 10,
            }
            result = ExcelLayoutService.expand_date_columns_kpis(
                worksheet=ws, sheet_definition="individual_kpis", ds="2024-10-01"
            )

        # Assert
        assert result is not None
        assert "date_mappings" in result
        assert "total_width" in result
        mock_expand.assert_called_once()

    def test_expand_date_columns_collective_delegates_correctly(self):
        """Test collective KPI column expansion delegates to internal method."""
        # Arrange
        wb = Workbook()
        ws = wb.active

        # Act - Mock the internal method
        with patch(
            "services.excel_layout.ExcelLayoutService._expand_date_columns_collective"
        ) as mock_expand:
            mock_expand.return_value = {
                "date_mappings": {"years": [], "months": []},
                "total_width": 8,
            }
            result = ExcelLayoutService.expand_date_columns_kpis(
                worksheet=ws, sheet_definition="collective_kpis", ds="2024-10-01"
            )

        # Assert
        assert result is not None
        assert "date_mappings" in result
        mock_expand.assert_called_once()

    def test_expand_date_columns_unknown_definition(self):
        """Test handling of unknown sheet definition."""
        # Arrange
        wb = Workbook()
        ws = wb.active

        # Act
        result = ExcelLayoutService.expand_date_columns_kpis(
            worksheet=ws, sheet_definition="unknown_kpis", ds="2024-10-01"
        )

        # Assert
        assert result is None


class TestSetSheetTitle:
    """Test cases for set_sheet_title method."""

    def test_set_sheet_title_kpis(self):
        """Test title setting for KPI sheets."""
        # Arrange
        wb = Workbook()
        ws = wb.active

        # Act
        with patch(
            "services.excel_layout.SHEET_LAYOUT",
            {
                "kpis": {
                    "title_row_offset": 0,
                    "title_col_offset": 0,
                    "title_height": 3,
                    "title_width": "dynamic",
                }
            },
        ):
            with patch(
                "services.excel_layout.SHEET_DEFINITIONS",
                {"individual_kpis": {"title_suffix": "part individuelle"}},
            ):
                ExcelLayoutService.set_sheet_title(
                    worksheet=ws,
                    title_base="Test Report",
                    layout_type="kpis",
                    context={"region": "Île-de-France"},
                    filters={"scale": "region"},
                    expanded_width=10,
                    sheet_definition="individual_kpis",
                )

        # Assert
        assert ws.cell(row=1, column=1).value is not None

    def test_set_sheet_title_top(self):
        """Test title setting for top sheets."""
        # Arrange
        wb = Workbook()
        ws = wb.active

        # Act
        with patch(
            "services.excel_layout.SHEET_LAYOUT",
            {
                "top": {
                    "title_row_offset": 0,
                    "title_col_offset": 0,
                    "title_height": 2,
                    "title_width": {"top_offers": 5},
                }
            },
        ):
            with patch(
                "services.excel_layout.SHEET_DEFINITIONS",
                {"top_offers": {"title_suffix": "classement"}},
            ):
                ExcelLayoutService.set_sheet_title(
                    worksheet=ws,
                    title_base="Top Offers",
                    layout_type="top",
                    context={},
                    filters={},
                    sheet_definition="top_offers",
                )

        # Assert
        assert ws.cell(row=1, column=1).value is not None


class TestCleanupTemplateColumns:
    """Test cases for cleanup_template_columns method."""

    def test_cleanup_template_columns_with_offset(self):
        """Test deletion of template columns."""
        # Arrange
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Template1")
        ws.cell(row=1, column=2, value="Template2")
        ws.cell(row=1, column=3, value="Data1")

        # Act
        with patch(
            "services.excel_layout.SHEET_LAYOUT", {"kpis": {"title_col_offset": 2}}
        ):
            ExcelLayoutService.cleanup_template_columns(ws, "kpis")

        # Assert
        # First two columns should be deleted, Data1 should now be in column 1
        assert ws.cell(row=1, column=1).value == "Data1"

    def test_cleanup_template_columns_no_offset(self):
        """Test when no template columns to delete."""
        # Arrange
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Data1")

        # Act
        with patch(
            "services.excel_layout.SHEET_LAYOUT", {"top": {"title_col_offset": 0}}
        ):
            ExcelLayoutService.cleanup_template_columns(ws, "top")

        # Assert
        assert ws.cell(row=1, column=1).value == "Data1"


class TestFreezePanes:
    """Test cases for freeze_panes method."""

    def test_freeze_panes_with_values(self):
        """Test setting freeze panes."""
        # Arrange
        wb = Workbook()
        ws = wb.active

        # Act
        with patch(
            "services.excel_layout.SHEET_LAYOUT",
            {"kpis": {"freeze_panes": {"row": 5, "col": 2}}},
        ):
            ExcelLayoutService.freeze_panes(ws, "kpis")

        # Assert
        assert ws.freeze_panes is not None
        assert (
            ws.freeze_panes == "C6"
        )  # col=2 -> C, row=5 -> 6 (0-indexed to 1-indexed)

    def test_freeze_panes_no_freeze(self):
        """Test when no freeze panes configured."""
        # Arrange
        wb = Workbook()
        ws = wb.active

        # Act
        with patch(
            "services.excel_layout.SHEET_LAYOUT",
            {"top": {"freeze_panes": {"row": 0, "col": 0}}},
        ):
            ExcelLayoutService.freeze_panes(ws, "top")

        # Assert - should not set freeze panes
        assert ws.freeze_panes is None or ws.freeze_panes == "A1"
