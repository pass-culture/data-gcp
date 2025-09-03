import openpyxl
from pathlib import Path
import duckdb
from typing import Dict, Any, List, Optional, TYPE_CHECKING
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import logging

# Use TYPE_CHECKING to avoid circular import
if TYPE_CHECKING:
    from core import Sheet, SheetType

logger = logging.getLogger(__name__)

class ExcelTemplateHandler:
    """Handles Excel template loading, data filling, and report generation."""
    
    def __init__(self, template_path: Path):
        self.template_path = template_path
        self.workbook: Optional[openpyxl.Workbook] = None
    
    def load_template(self) -> openpyxl.Workbook:
        """Load the Excel template."""
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template not found: {self.template_path}")
        
        self.workbook = openpyxl.load_workbook(self.template_path)
        logger.info(f"Loaded template: {self.template_path}")
        return self.workbook
    
    def create_or_get_worksheet(self, sheet_name: str) -> openpyxl.worksheet.worksheet.Worksheet:
        """Create a new worksheet or get existing one."""
        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        else:
            return self.workbook.create_sheet(sheet_name)
    
    def fill_sheet_with_data(self, sheet: "Sheet", conn: duckdb.DuckDBPyConnection, filters: Dict[str, str]):
        """Fill a sheet with data from DuckDB."""
        try:
            # Get or create worksheet
            ws = self.create_or_get_worksheet(sheet.name)
            
            # Import here to avoid circular import
            from core import SheetType
            
            if sheet.type == SheetType.LEXIQUE:
                self._fill_lexique_sheet(ws)
                return
            
            # Build and execute query
            query = sheet.build_query(filters)
            logger.info(f"Executing query for sheet {sheet.name}: {query[:100]}...")
            
            df = conn.execute(query).df()
            
            for col in df.select_dtypes(include=['datetime64[ns, UTC]', 'datetimetz']).columns:
                    df[col] = df[col].dt.tz_localize(None)
                    
            if df.empty:
                logger.warning(f"No data found for sheet {sheet.name}")
                return
            
            # Fill data based on sheet type
            if sheet.type == SheetType.KPIS:
                self._fill_kpi_sheet(ws, df, sheet)
            elif sheet.type == SheetType.TOP:
                self._fill_top_sheet(ws, df, sheet)
                
            logger.info(f"Filled sheet {sheet.name} with {len(df)} rows")
            
        except Exception as e:
            logger.error(f"Error filling sheet {sheet.name}: {e}")
            raise
    
    def _fill_kpi_sheet(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                        df: pd.DataFrame, sheet: "Sheet"):
        """Fill KPI sheet with data."""
        # Clear existing data (preserve headers if any)
        self._clear_data_area(worksheet)
        
        # Add headers
        headers = ['Date', 'Dimension', 'Valeur', 'KPI', 'Numérateur', 'Dénominateur', 'Résultat']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add data rows
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            worksheet.cell(row=row_idx, column=1).value = row.get('partition_month')
            worksheet.cell(row=row_idx, column=2).value = row.get('dimension_name')
            worksheet.cell(row=row_idx, column=3).value = row.get('dimension_value')
            worksheet.cell(row=row_idx, column=4).value = row.get('kpi_name')
            worksheet.cell(row=row_idx, column=5).value = row.get('numerator')
            worksheet.cell(row=row_idx, column=6).value = row.get('denominator')
            worksheet.cell(row=row_idx, column=7).value = row.get('kpi')
        
        # Apply formatting
        self._apply_table_formatting(worksheet, len(df) + 1, len(headers))
    
    def _fill_top_sheet(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                       df: pd.DataFrame, sheet: "Sheet"):
        """Fill TOP sheets with data."""
        # Clear existing data
        self._clear_data_area(worksheet)
        
        # Get all columns from dataframe
        headers = list(df.columns)
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add data rows
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, header in enumerate(headers, 1):
                worksheet.cell(row=row_idx, column=col_idx).value = row[header]
        
        # Apply formatting
        self._apply_table_formatting(worksheet, len(df) + 1, len(headers))
    
    def _fill_lexique_sheet(self, worksheet: openpyxl.worksheet.worksheet.Worksheet):
        """Fill lexique/glossary sheet with standard definitions."""
        # Clear existing content
        worksheet.delete_rows(1, worksheet.max_row)
        
        # Add title
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = "Lexique des Indicateurs"
        title_cell.font = Font(bold=True, size=16)
        
        # Add definitions (customize as needed)
        definitions = [
            ("KPI", "Indicateur Clé de Performance"),
            ("Numérateur", "Valeur du numérateur pour le calcul du KPI"),
            ("Dénominateur", "Valeur du dénominateur pour le calcul du KPI"),
            ("Dimension", "Niveau géographique (NAT, REG, DEP, ACA)"),
            ("Part individuelle", "Données relatives aux offres individuelles"),
            ("Part collective", "Données relatives aux offres collectives"),
        ]
        
        # Headers for definitions
        worksheet.cell(row=3, column=1).value = "Terme"
        worksheet.cell(row=3, column=2).value = "Définition"
        
        for row_idx, (term, definition) in enumerate(definitions, 4):
            worksheet.cell(row=row_idx, column=1).value = term
            worksheet.cell(row=row_idx, column=2).value = definition
        
        # Apply formatting
        for row in worksheet.iter_rows(min_row=3, max_row=3, min_col=1, max_col=2):
            for cell in row:
                cell.font = Font(bold=True)
        
        # Adjust column widths
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 50
    
    def _clear_data_area(self, worksheet: openpyxl.worksheet.worksheet.Worksheet):
        """Clear data area while preserving structure."""
        if worksheet.max_row > 1:
            worksheet.delete_rows(2, worksheet.max_row - 1)
    
    def _apply_table_formatting(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                               max_row: int, max_col: int):
        """Apply basic table formatting."""
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = thin_border
        
        # Alternate row colors for better readability
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        for row_num in range(3, max_row + 1, 2):  # Every other row starting from row 3
            for col_num in range(1, max_col + 1):
                worksheet.cell(row=row_num, column=col_num).fill = light_fill
        
        # Auto-adjust column widths
        for col_num in range(1, max_col + 1):
            column_letter = openpyxl.utils.get_column_letter(col_num)
            max_length = 0
            for row_num in range(1, min(max_row + 1, 50)):  # Check first 50 rows for performance
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def organize_worksheets(self):
        """Organize worksheet order - move Lexique to front."""
        if not self.workbook:
            return
            
        # Move Lexique sheet to the beginning if it exists
        if "Lexique" in self.workbook.sheetnames:
            lexique_sheet = self.workbook["Lexique"]
            self.workbook.move_sheet(lexique_sheet, offset=-len(self.workbook.sheetnames) + 1)
        
        # Remove any empty default sheets
        sheets_to_remove = []
        for sheet in self.workbook.worksheets:
            if sheet.title in ['Sheet', 'Sheet1', 'Feuil1'] and sheet.max_row <= 1:
                sheets_to_remove.append(sheet)
        
        for sheet in sheets_to_remove:
            self.workbook.remove(sheet)
    
    def save_report(self, output_path: Path):
        """Save the filled workbook."""
        if not self.workbook:
            raise ValueError("No workbook loaded")
        
        # Organize sheets before saving
        self.organize_worksheets()
        
        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Save the workbook
        self.workbook.save(output_path)
        logger.info(f"Saved report: {output_path}")
    
    def close(self):
        """Clean up resources."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None