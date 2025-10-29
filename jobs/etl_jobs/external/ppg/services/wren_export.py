"""
Export Excel template data to text files for Wren AI knowledge base.

This module extracts structured data from the export_template.xlsx file
and creates text files organized by logical sections, with each file
containing up to 1000 characters of related content.
"""

from pathlib import Path
from typing import List

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class WrenExporter:
    """Export Excel template data to text files for Wren AI."""

    def __init__(self, template_path: Path, output_dir: Path, max_chars: int = 1000):
        """
        Initialize the Wren exporter.

        Args:
            template_path: Path to the Excel template file
            output_dir: Directory where text files will be created
            max_chars: Maximum characters per text file (default: 1000)
        """
        self.template_path = template_path
        self.output_dir = output_dir
        self.max_chars = max_chars
        self.wb = None

    def load_workbook(self):
        """Load the Excel workbook."""
        self.wb = openpyxl.load_workbook(self.template_path, data_only=True)

    def export_all(self):
        """Export all sheets to text files."""
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.load_workbook()

        # Export each sheet with its specific logic
        self._export_lexique()
        self._export_template_sheet("template_individuel")
        self._export_template_sheet("template_collectif")

        print(f"✅ Export completed. Files saved in: {self.output_dir}")

    def _export_lexique(self):
        """Export Lexique sheet by grouping indicators and their descriptions."""
        ws = self.wb["Lexique"]

        # Parse the sheet to find indicator sections
        sections = self._parse_lexique_sections(ws)

        # Write sections to files
        for section_name, content_items in sections.items():
            self._write_chunked_files(section_name, content_items)

    def _parse_lexique_sections(self, ws: Worksheet) -> dict:
        """
        Parse Lexique sheet into logical sections.

        Returns:
            Dictionary mapping section names to lists of content strings
        """
        sections = {}
        current_section = None

        for row in ws.iter_rows(values_only=True):
            # Get first two columns (indicator name and description)
            col_a = str(row[0]).strip() if row[0] else ""
            col_b = str(row[1]).strip() if row[1] else ""

            # Skip empty rows
            if not col_a and not col_b:
                continue

            # Detect section headers (single column text, no description)
            if col_a and not col_b and len(col_a) < 50:
                # This is likely a section header
                current_section = self._slugify(col_a)
                if current_section not in sections:
                    sections[current_section] = []
            elif col_a and col_b:
                # This is an indicator with description
                if current_section is None:
                    current_section = "general"
                    sections[current_section] = []

                content = f"{col_a}: {col_b}"
                sections[current_section].append(content)
            elif col_a:
                # Single column with substantial text
                if current_section is None:
                    current_section = "general"
                    sections[current_section] = []
                sections[current_section].append(col_a)

        return sections

    def _export_template_sheet(self, sheet_name: str):
        """
        Export template sheet (individuel or collectif) with exact kpi_name mappings.

        Args:
            sheet_name: Name of the sheet to export
        """
        ws = self.wb[sheet_name]

        # Parse KPI definitions (kpi_name -> definition string)
        kpi_definitions = self._parse_template_kpis(ws)

        # Write each KPI to its own file
        prefix = "individuel" if "individuel" in sheet_name else "collectif"
        for kpi_name_exact, definition_content in kpi_definitions.items():
            # Slugify for filename but keep exact name in content
            filename_base = f"{prefix}_{self._slugify(kpi_name_exact)}"
            self._write_single_file(filename_base, definition_content)

    def _parse_template_kpis(self, ws: Worksheet) -> dict:
        """
        Parse template sheet into KPI groups, preserving exact kpi_name.

        Returns:
            Dictionary mapping exact KPI names to their definitions
        """
        kpi_definitions = {}
        current_section_header = None

        for row in ws.iter_rows(values_only=True):
            # Get relevant columns
            col_filter = str(row[0]).strip() if row[0] else ""
            col_select = str(row[1]).strip() if row[1] else ""
            col_calcul = str(row[2]).strip() if row[2] else ""
            col_label = str(row[3]).strip() if row[3] else ""

            # Skip empty rows
            if not any([col_filter, col_select, col_calcul, col_label]):
                continue

            # Detect section headers (descriptive text in col_label, no filter)
            if (
                col_label
                and not col_filter.startswith("kpi_name")
                and len(col_label) > 15
            ):
                if not any(
                    col_label.startswith(prefix)
                    for prefix in ["YYYY", "Par an", "A décliner"]
                ):
                    current_section_header = col_label

            # Detect KPI definitions
            if col_filter.startswith("kpi_name"):
                # Extract EXACT KPI name from filter column (no slugification)
                kpi_name_exact = (
                    col_filter.replace("kpi_name =", "")
                    .replace("kpi_name=", "")
                    .strip()
                )

                if not kpi_name_exact:
                    continue

                # Create KPI definition with clear mapping format
                definition_parts = [f"kpi_name: {kpi_name_exact}"]

                if col_label:
                    definition_parts.append(f"Definition: {col_label}")
                if col_select:
                    definition_parts.append(f"Type: {col_select}")
                if col_calcul:
                    definition_parts.append(f"Calculation: {col_calcul}")
                if current_section_header:
                    definition_parts.append(f"Section: {current_section_header}")

                # Use exact kpi_name as key (one definition per KPI)
                kpi_content = "\n".join(definition_parts)
                kpi_definitions[kpi_name_exact] = kpi_content

        return kpi_definitions

    def _write_chunked_files(self, section_name: str, content_items: List[str]):
        """
        Write content items to files, respecting max_chars limit.

        Args:
            section_name: Base name for the files
            content_items: List of content strings to write
        """
        if not content_items:
            return

        chunk_idx = 0
        current_chunk = []
        current_length = 0

        for item in content_items:
            item_length = len(item)

            # If single item exceeds max_chars, write it alone
            if item_length > self.max_chars:
                if current_chunk:
                    self._write_file(section_name, chunk_idx, current_chunk)
                    chunk_idx += 1
                    current_chunk = []
                    current_length = 0

                self._write_file(section_name, chunk_idx, [item])
                chunk_idx += 1
                continue

            # If adding this item would exceed limit, write current chunk
            if current_length + item_length + 2 > self.max_chars:  # +2 for newlines
                self._write_file(section_name, chunk_idx, current_chunk)
                chunk_idx += 1
                current_chunk = []
                current_length = 0

            # Add item to current chunk
            current_chunk.append(item)
            current_length += item_length + 2  # +2 for "\n\n"

        # Write remaining chunk
        if current_chunk:
            self._write_file(section_name, chunk_idx, current_chunk)

    def _write_file(self, section_name: str, chunk_idx: int, content_items: List[str]):
        """
        Write a single text file with chunked content.

        Args:
            section_name: Base name for the file
            chunk_idx: Chunk index
            content_items: Content to write
        """
        filename = f"{section_name}_{chunk_idx}.txt"
        filepath = self.output_dir / filename

        content = "\n\n".join(content_items)
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(content)

        print(f"  ✓ Created: {filename} ({len(content)} chars)")

    def _write_single_file(self, filename_base: str, content: str):
        """
        Write a single KPI definition file without chunking.

        Args:
            filename_base: Base name for the file (without .txt extension)
            content: Content to write
        """
        filename = f"{filename_base}.txt"
        filepath = self.output_dir / filename

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(content)

        print(f"  ✓ Created: {filename} ({len(content)} chars)")

    @staticmethod
    def _slugify(text: str) -> str:
        """
        Convert text to a filename-safe slug.

        Args:
            text: Text to slugify

        Returns:
            Slugified string
        """
        # Remove common prefixes
        text = text.strip()
        for prefix in [" ", "kpi_name =", "kpi_name=", "Filter", "Select"]:
            text = text.replace(prefix, "")

        # Convert to lowercase and replace spaces/special chars
        slug = text.lower()
        slug = slug.replace(" ", "_")
        slug = slug.replace("'", "")
        slug = slug.replace("é", "e")
        slug = slug.replace("è", "e")
        slug = slug.replace("ê", "e")
        slug = slug.replace("à", "a")
        slug = slug.replace("ô", "o")
        slug = slug.replace("ù", "u")
        slug = slug.replace("î", "i")
        slug = slug.replace("ç", "c")

        # Remove non-alphanumeric characters (except underscore and hyphen)
        slug = "".join(c for c in slug if c.isalnum() or c in ["_", "-"])

        # Remove consecutive underscores
        while "__" in slug:
            slug = slug.replace("__", "_")

        slug = slug.strip("_-")

        return slug[:100]  # Limit length (increased to avoid truncation)


def export_to_wren(
    template_path: str = "templates/export_template.xlsx",
    output_dir: str = "wren_knowledge",
    max_chars: int = 1000,
) -> None:
    """
    Export Excel template to Wren AI knowledge base text files.

    Args:
        template_path: Path to the Excel template file
        output_dir: Directory where text files will be created
        max_chars: Maximum characters per text file
    """
    template_path = Path(template_path)
    output_dir = Path(output_dir)

    if not template_path.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")

    exporter = WrenExporter(template_path, output_dir, max_chars)
    exporter.export_all()


if __name__ == "__main__":
    export_to_wren()
