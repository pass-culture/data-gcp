# models.py
from dataclasses import dataclass
from pathlib import Path
from typing import List, Dict, Optional, Any
import duckdb
import openpyxl

@dataclass
class TargetStakeholder:
    """Represents who gets a report and their filtering requirements."""
    name: str   # Ministrere, DRAC
    dimension_name: str  # NAT, REG
    dimension_value: str
    level: str  # national, regional, academie, departemental
    scope: str  # individual, collective
    region_name: Optional[str] = None
    academie_name: Optional[str] = None
    
    def get_data_filters(self) -> Dict[str, str]:
        """Generate SQL WHERE clause filters."""
        filters = {"dimension_name": self.dimension_name}
        if self.dimension_value != self.dimension_name:  # Not national
            filters["dimension_value"] = self.dimension_value
        return filters
    
    def get_output_filename(self) -> str:
        """Generate appropriate filename."""
        if self.level == "national":
            return "rapport_national.xlsx"
        elif self.level == "regional":
            return f"rapport_{self.dimension_value}.xlsx"
        elif self.level == "academie":
            return f"collectif_academie_{self.dimension_value}.xlsx"
        elif self.level == "departemental":
            return f"individuel_{self.dimension_value}.xlsx"
        return f"report_{self.name}.xlsx"

@dataclass
class Sheet:
    """Represents a sheet within a report."""
    sheet_type: str  # individuel, collectif, tops_individuel, tops_collectif
    worksheet: openpyxl.worksheet.worksheet.Worksheet
    kpi_filters: List[tuple]  # (row_idx, kpi_name, select_type, agg_type)
    
    def fill_with_data(self, conn: duckdb.DuckDBPyConnection, stakeholder: TargetStakeholder, ds: str):
        """Fill this sheet with data for the stakeholder."""
        # Move your existing sheet filling logic here
        pass
    
    def _apply_formatting(self):
        """Apply formatting to the worksheet."""
        pass

@dataclass
class Report:
    """Represents a complete report file."""
    stakeholder: TargetStakeholder
    base_template_path: Path
    output_path: Path
    scope: str = "all"   # individual, collective, all
    sheets: List[Sheet] = None,
    workbook: Optional[openpyxl.Workbook] = None
    
    def load_workbook_template(self):
        """Load the Excel workbook."""
        self.workbook = openpyxl.load_workbook(self.template_path)
    
    def generate_sheets(self):
        """Initialize sheets based on the template."""
        if not self.workbook:
            self.load_workbook()
        
        
        self.sheets = []
        
        for sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]
            # Determine sheet_type and kpi_filters based on sheet_name or content
            sheet_type = "individual"
    
    def generate(self, conn: duckdb.DuckDBPyConnection, ds: str):
        """Generate the complete report."""
        if not self.workbook:
            self.load_workbook()
        
        # Fill each sheet
        for sheet in self.sheets:
            sheet.fill_with_data(conn, self.stakeholder, ds)
        
        # Apply final formatting and save
        self._finalize_and_save()
    
    def _finalize_and_save(self):
        """Apply final formatting and save the report."""
        # Move lexique sheet first, hide columns, etc.
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(self.output_path)
