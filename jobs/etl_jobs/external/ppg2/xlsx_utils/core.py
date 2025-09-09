from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Dict, Optional, Any
from enum import Enum
from collections import defaultdict
from duckdb import DuckDBPyConnection

from treelib import Tree
import typer
import openpyxl
from config import (REGION_HIERARCHY_TABLE,GCP_PROJECT, BIGQUERY_ANALYTICS_DATASET, ENV_SHORT_NAME,)
from google.cloud import bigquery

from utils.file_utils import slugify

import os
from copy import copy
from datetime import datetime, timedelta,date
from dateutil.relativedelta import relativedelta
import pandas as pd
from utils.duckdb_utils import aggregate_kpi_data, query_yearly_kpi, query_monthly_kpi

def build_region_hierarchy(
    project_id: str = GCP_PROJECT,
    dataset: str = BIGQUERY_ANALYTICS_DATASET,
    table: str = REGION_HIERARCHY_TABLE,
) -> Dict[str, Dict]:
    """
    Build a nested dict with enhanced academy-department mapping from BigQuery.

    Args:
        project_id: GCP project ID
        dataset: BigQuery dataset
        table: BigQuery table name

    Returns:
        Dict with structure:
        {
            region_name: {
                "departements": [dept1, dept2, ...],
                "academies": [acad1, acad2, ...],
                "academy_departments": {
                    academy_name: [dept1, dept2, ...]
                }
            }
        }

    Raises:
        HierarchyError: If hierarchy building fails
    """

    try:
        client = bigquery.Client(project=project_id)

        query = f"""
        SELECT
            region_name,
            dep_name,
            academy_name
        FROM `{project_id}.{dataset}.{table}`
        WHERE region_name IS NOT NULL
        ORDER BY region_name, academy_name, dep_name
        """

        typer.echo(f"🌍 Querying region hierarchy: {project_id}.{dataset}.{table}")

        rows = client.query(query).result()
        row_count = 0
        by_region = {}

        for row in rows:
            row_count += 1
            region = row.region_name
            dep = row.dep_name
            acad = row.academy_name

            if region not in by_region:
                by_region[region] = {
                    "departements": set(),
                    "academies": set(),
                    "academy_departments": {},
                }

            bucket = by_region[region]

            # Add department if present
            if dep:
                bucket["departements"].add(dep)

            # Add academy if present
            if acad:
                bucket["academies"].add(acad)

            # Map departments to academies
            if acad and dep:
                if acad not in bucket["academy_departments"]:
                    bucket["academy_departments"][acad] = set()
                bucket["academy_departments"][acad].add(dep)

        # Convert sets to sorted lists for consistent output
        sorted_by_region = {}
        for region, scope in sorted(by_region.items()):
            sorted_by_region[region] = {
                "departements": sorted(scope["departements"]),
                "academies": sorted(scope["academies"]),
                "academy_departments": {
                    academy: sorted(departments)
                    for academy, departments in scope["academy_departments"].items()
                },
            }

        # Log summary statistics
        total_regions = len(sorted_by_region)
        total_academies = sum(len(r["academies"]) for r in sorted_by_region.values())
        total_departments = sum(
            len(r["departements"]) for r in sorted_by_region.values()
        )

        typer.echo(
            f"✅ Loaded hierarchy: {total_regions} regions, {total_academies} academies, {total_departments} departments"
        )
        typer.echo(f"   📊 Total records processed: {row_count:,}")

        return sorted_by_region
    except Exception as e:
        typer.secho(f"❌ Failed to build region hierarchy: {e}", fg="red")
        raise


######## base configs
# GCP_PROJECT = os.environ.get("PROJECT_NAME", "passculture-data-prod")
# ENV_SHORT_NAME = os.environ.get("ENV_SHORT_NAME", "prod")
# BIGQUERY_ANALYTICS_DATASET = f"analytics_{ENV_SHORT_NAME}"


# REGION_HIERARCHY_TABLE = "region_department"

# TAB_NAME_MAX_LENGTH = 31

# ALL_TABLES = {
#     "individual": {"table_name": f"{table_prefix}individual", "tab": "Part individuelle", "scope": "individual"},
#     "collective": {"table_name": f"{table_prefix}collective", "tab": "Part collective", "scope": "collective"},
#     "top_individual_offer": {"table_name": f"{table_prefix}top_offer", "tab": "TOP 50 offres - individuel", "scope":"individual"},
#     "top_individual_offer_category": {"table_name": f"{table_prefix}top_offer_category", "tab": "TOP 50 offres par catégorie - individuel", "scope":"individual"},
#     "top_individual_venue": {"table_name": f"{table_prefix}top_venue", "tab": "TOP 50 lieux - individuel", "scope":"individual"},
# }




STAKEHOLDER_REPORTS = {
    "ministere": ["national_summary"],
    "drac": ["regional_summary", "departemental_detail", "academy_detail"]
}

# BASE_TEMPLATE = Path("./jobs/etl_jobs/external/ppg/templates/PPG_Template_v4.xlsx")
BASE_TEMPLATE = Path("./templates/export_template_v4.xlsx")

TEMPLATE_DEFAULT = Path("./templates/export_template_v2.xlsx")



# data sources
table_prefix = "external_reporting"
# One Big Table (kpi per scope)
# KPI_INDIV = {"dataset": BIGQUERY_ANALYTICS_DATASET, "table": f"{table_prefix}individual"}
# KPI_COLL = {"dataset": BIGQUERY_ANALYTICS_DATASET, "table": f"{table_prefix}eac"}

# TOP_INDIV_OFFER = {"dataset": BIGQUERY_ANALYTICS_DATASET, "table": f"{table_prefix}top_offer"}
# TOP_INDIV_OFFER_CAT = {"dataset": BIGQUERY_ANALYTICS_DATASET, "table": f"{table_prefix}top_offer_category"}
# TOP_INDIV_VENUE = {"dataset": BIGQUERY_ANALYTICS_DATASET, "table": f"{table_prefix}top_venue"}


# SHEETS = {
#     "Lexique": {
#         "template": TEMPLATE_DEFAULT,
#         "tab": "Lexique",
#         "data": None,
#         },
#     "Individual_KPIs": {
#         "template": TEMPLATE_DEFAULT,
#         "tab": "template_individuel",
#         "data": KPI_INDIV,
#         },
#     "Collective_KPIs": {
#         "template": TEMPLATE_DEFAULT,
#         "tab": "template_collective",
#         "data": KPI_COLL,
#         },
#     "Top_individual_Offer": {
#         "template": TEMPLATE_DEFAULT,
#         "tab": "template_top",
#         "data": TOP_INDIV_OFFER,
#         "headers": None
#         },
#     "Top_individual_Offer_Category": {
#         "template": TEMPLATE_DEFAULT,
#         "tab": "template_top",
#         "data": TOP_INDIV_OFFER_CAT,
#         },
#     "Top_individual_Venue": {
#         "template": TEMPLATE_DEFAULT,
#         "tab": "template_top",
#         "data": TOP_INDIV_VENUE,
#         },
# }

SOURCE_TABLES = {
    "individual": {"dataset": BIGQUERY_ANALYTICS_DATASET, "table": f"{table_prefix}_individual"},
    "collective": {"dataset": BIGQUERY_ANALYTICS_DATASET, "table": f"{table_prefix}_eac"},
    "top_offer": {"dataset": BIGQUERY_ANALYTICS_DATASET, "table": f"{table_prefix}_top_offer"},
    "top_offer_category": {"dataset": BIGQUERY_ANALYTICS_DATASET, "table": f"{table_prefix}_top_offer_category"},
    "top_venue": {"dataset": BIGQUERY_ANALYTICS_DATASET, "table": f"{table_prefix}_top_venue"},
    "top_ac": {"dataset": BIGQUERY_ANALYTICS_DATASET, "table": f"{table_prefix}_top_ac"},
    "top_format": {"dataset": BIGQUERY_ANALYTICS_DATASET, "table": f"{table_prefix}_top_format"},
}
class SheetType(Enum):
    KPIS = "kpis"
    TOP = "top"
    LEXIQUE = "lexique"

SHEET_DEFINITIONS = {
    "individual_kpis": {
        "type": SheetType.KPIS,
        "template_tab": "template_individuel",
        "source_table": "individual",
    },
    "collective_kpis": {
        "type": SheetType.KPIS,
        "template_tab": "template_collectif",
        "source_table": "collective",
    },
    "lexique": {
        "type": SheetType.LEXIQUE,
        "template_tab": "Lexique",
        "source_table": None,
    },
    "top_offer": {
        "type": SheetType.TOP,
        "template_tab": "Top 50 offres",
        "source_table": "top_offer",
        "top_n": 50,
    },
    "top_offer_category": {
        "type": SheetType.TOP,
        "template_tab": "Top 50 par catégorie",
        "source_table": "top_offer_category",
        "top_n": 50,
    },
    "top_venue": {
        "type": SheetType.TOP,
        "template_tab": "Top 50 lieux",
        "source_table": "top_venue",
        "top_n": 50,
    },
    "top_ac": {
        "type": SheetType.TOP,
        "template_tab": "Top 50 acteurs culturels",
        "source_table": "top_ac",
        "top_n": 50,
    },
    "top_format": {
        "type": SheetType.TOP,
        "template_tab": "Top 5 formats",
        "source_table": "top_format",
        "top_n": 5,
    },
}

def default_layout():
    return {"title_row_offset": 0, "title_col_offset": 0}

SHEET_LAYOUT = defaultdict(default_layout, {
    "top": {"title_row_offset": 0, "title_col_offset": 0},
    "kpis": {"title_row_offset": 0, "title_col_offset": 3}
    }
)

REPORTS = {
    "national_summary": {
        "sheets": [
            {"definition": "lexique"},
            {"definition": "individual_kpis", "filters": {"scope": "individual", "scale": "national"}},
            {"definition": "top_offer", "filters": {"scope": "individual", "scale": "national"}},
            {"definition": "top_offer_category", "filters": {"scope": "individual", "scale": "national"}},
            {"definition": "top_venue", "filters": {"scope": "individual", "scale": "national"}},
            {"definition": "collective_kpis", "filters": {"scope": "collective", "scale": "national"}},
            {"definition": "top_ac", "filters": {"scope": "collective", "scale": "national"}},
            {"definition": "top_format", "filters": {"scope": "collective", "scale": "national"}},
        ]
    },
    "region_summary": {
        "sheets": [
            {"definition": "lexique"},
            {"definition": "individual_kpis", "filters": {"scope": "individual", "scale": "national"}},
            {"definition": "individual_kpis", "filters": {"scope": "individual", "scale": "region"}},
            {"definition": "top_offer", "filters": {"scope": "individual", "scale": "region"}},
            {"definition": "top_offer_category", "filters": {"scope": "individual", "scale": "region"}},
            {"definition": "top_venue", "filters": {"scope": "individual", "scale": "region"}},
            {"definition": "collective_kpis", "filters": {"scope": "collective", "scale": "national"}},
            {"definition": "collective_kpis", "filters": {"scope": "collective", "scale": "region"}},
            {"definition": "top_ac", "filters": {"scope": "collective", "scale": "region"}},
            {"definition": "top_format", "filters": {"scope": "collective", "scale": "region"}},
        ]
    },
    "academy_detail": {
        "sheets": [
            {"definition": "lexique"},
            {"definition": "collective_kpis", "filters": {"scope": "collective", "scale": "region"}},
            {"definition": "collective_kpis", "filters": {"scope": "collective", "scale": "academy"}},
            {"definition": "top_ac", "filters": {"scope": "collective", "scale": "academy"}},
            {"definition": "top_format", "filters": {"scope": "collective", "scale": "academy"}},
        ]
    },
    "department_detail": {
        "sheets": [
            {"definition": "lexique"},
            {"definition": "individual_kpis", "filters": {"scope": "individual", "scale": "region"}},
            {"definition": "individual_kpis", "filters": {"scope": "individual", "scale": "department"}},
            {"definition": "top_offer", "filters": {"scope": "individual", "scale": "department"}},
            {"definition": "top_offer_category", "filters": {"scope": "individual", "scale": "department"}},
            {"definition": "top_venue", "filters": {"scope": "individual", "scale": "department"}},
        ]
    },
}

# BASIC_QUERY_TEMPLATE = """
# SELECT partition_month, updated_at, dimension_name, dimension_value,
#        kpi_name, numerator, denominator, kpi
# FROM `{project}.{dataset}.{table}`
# WHERE 1=1
# ORDER BY partition_month, dimension_name, dimension_value, kpi_name
# """

# TOP_QUERY_TEMPLATE = """
# SELECT *
# FROM `{project}.{dataset}.{table}` 
# WHERE 1=1
# ORDER BY partition_month, dimension_name, dimension_value
# """


class StakeholderType(Enum):
    MINISTERE = "ministere"
    DRAC = "drac"


@dataclass
class Stakeholder:
    """Stakeholder - either Ministère (no children) or DRAC (region → academies & departments)."""
    type: StakeholderType
    name: str
    desired_reports: List[str] = field(default_factory=list)
    academy_tree: Optional[Tree] = None
    department_tree: Optional[Tree] = None

    def __post_init__(self):
        # Assign default reports
        if not self.desired_reports:
            self.desired_reports = list(STAKEHOLDER_REPORTS.get(self.type.value, []))

        # Only DRAC stakeholders get trees
        if self.type == StakeholderType.DRAC:
            # Fetch hierarchy from BigQuery
            hierarchy = build_region_hierarchy()

            # Ensure region exists
            if self.name not in hierarchy:
                raise ValueError(f"❌ Region '{self.name}' not found in hierarchy")

            scope = hierarchy[self.name]

            # Build departement tree
            self.department_tree = Tree()
            self.department_tree.create_node("Departements", "root_dept")
            for dep in scope["departements"]:
                self.department_tree.create_node(dep, f"{self.name}_dep_{dep}", parent="root_dept")

            # Build academy tree
            self.academy_tree = Tree()
            self.academy_tree.create_node("Academies", "root_acad")
            for acad in scope["academies"]:
                self.academy_tree.create_node(acad, f"{self.name}_acad_{acad}", parent="root_acad")

        else:
            self.academy_tree = None
            self.department_tree = None
        if self.type == StakeholderType.DRAC:
            typer.secho(f"➡️ DRAC '{self.name}' with adademy tree:\n{self.academy_tree.show(stdout=False)}", fg="cyan")
            typer.secho(f"➡️ DRAC '{self.name}' with department tree:\n{self.department_tree.show(stdout=False)}", fg="cyan")

@dataclass
class Sheet:
    """Represents a sheet instance within a report."""
    definition: str                    # key from SHEET_DEFINITIONS
    template_tab: str                  # template tab name
    worksheet: Any                     # openpyxl worksheet instance
    filters: Dict[str, Any] = field(default_factory=dict)
    tab_name: Optional[str] = None
    stakeholder: Optional[Any] = None  # reference to Stakeholder
    context: Optional[Dict[str, Any]] = None  # region, academy, department, etc.
    layout: Optional[Dict[str, Any]] = None

    def set_title(self):
        """Insert title in the worksheet according to SHEET_LAYOUT and rules."""
        layout_type = "top" if self.definition.startswith("top") else "kpis" if self.definition.endswith("kpis") else "other"
        if not self.layout:
            self.layout = SHEET_LAYOUT[layout_type]
        row = self.layout.get("title_row_offset")
        col = self.layout.get("title_col_offset")

        # Build title
        title_base = self.tab_name or self.definition.capitalize()
        
        #
        if layout_type in ["kpis","top"]:
            scale = self.filters.get("scale", "")
            node_tag = self.context.get(scale)
            title = title_base
            if node_tag:
                title += "\n" + node_tag.capitalize()
        else:
            title = title_base

        # Insert into worksheet
        self.worksheet.cell(row=row + 1, column=col + 1, value=title)
        
    def preprocess(self,ds:str):
        """Placeholder for preprocessing formatting, column widths, titles."""
        is_top = self.definition.startswith("top")
        is_kpis = self.definition.endswith("kpis")
        if (is_top ^ is_kpis) or not (is_top and is_kpis):
            self.set_title()
        else:
            raise ValueError(f"Sheet definition: {self.definition} cannot start with 'top' and end with 'kpis'")
        if is_kpis:
            self._expand_date_columns_kpis(ds)
            
    
    def _expand_date_columns_indiv(self, ds: str,min_year: int = 2021, nblank_cols: int = 1):
        """
        Expand columns in individual_kpis sheet based on ds (YYYY-MM).
        Copies template columns fully and updates header/date cells.
        Includes blank column and subsequent template columns (e.g., months).
        """

        if self.definition != "individual_kpis":
            return

        ds_year, ds_month, _ = map(int, ds.split("-"))
        ws = self.worksheet

        # Layout info
        layout = SHEET_LAYOUT.get(self.definition, {}).get("layout", {})
        title_row_offset = layout.get("title_row", 0)
        title_height = layout.get("title_height", 3)
        start_row = title_row_offset + title_height + 1
        header_row = ws[start_row]

        # Locate template "YYYY" column
        template_col_idx = None
        for idx, cell in enumerate(header_row[:10], start=1):
            if cell.value == "YYYY":
                template_col_idx = idx
                break
        if template_col_idx is None:
            raise ValueError("No 'YYYY' template column found in first 10 columns of the KPI sheet.")
        
        month_shifts = [-2,-1,-13]
        insert_idx = template_col_idx + nblank_cols + len(month_shifts) +1
        
        # Build mappings
        years_mapping, months_mapping = [], []
        mapping_offset = nblank_cols + len(month_shifts)+1
        
        # 1) Insert past years until ds_year-1
        for year in range(min_year, ds_year):
            ws.insert_cols(insert_idx)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                src_cell = row[template_col_idx - 1]
                dest_cell = row[insert_idx - 1]
                if src_cell.has_style:
                    dest_cell._style = copy(src_cell._style)
            label = f"{year}"
            ws.cell(row=start_row, column=insert_idx, value=label)
            years_mapping.append({int(insert_idx-mapping_offset): label})
            insert_idx += 1
            
        
        # 2) Insert blank columns
        for i in range(nblank_cols):
            ws.insert_cols(insert_idx)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                src_cell = row[template_col_idx + i]
                dest_cell = row[insert_idx - 1]
                if src_cell.has_style:
                    dest_cell._style = copy(src_cell._style)
            insert_idx += 1
            
        # 3) Insert month columns from template
        for month in month_shifts:
            target_date = datetime(ds_year, ds_month, 1) + relativedelta(months=month)
            ws.insert_cols(insert_idx)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                src_cell = row[template_col_idx - 1]
                dest_cell = row[insert_idx - 1]
                if src_cell.has_style:
                    dest_cell._style = copy(src_cell._style)
            label = target_date.strftime("%m/%Y")
            ws.cell(row=start_row, column=insert_idx, value=label)
            months_mapping.append({int(insert_idx-mapping_offset): label})
            insert_idx += 1
            
        # 4) Remove template columns
        ws.delete_cols(template_col_idx,nblank_cols + len(month_shifts)+1)
        
        layout["date_mappings"] = {"years": years_mapping, "months": months_mapping}
        self.layout = layout
            
    def _expand_date_columns_collective(self, ds: str,min_year: int = 2021, nblank_cols: int = 1):
        """
        Expand columns in collective_kpis sheet based on ds (YYYY-MM).
        Copies template columns fully and only updates header/date cells.
        """
        
        if self.definition != "collective_kpis":
            return

        ds_year, ds_month, _ = map(int, ds.split("-"))
        ws = self.worksheet

        # Layout info
        layout = SHEET_LAYOUT.get(self.definition, {}).get("layout", {})
        title_row_offset = layout.get("title_row", 0)
        title_height = layout.get("title_height", 3)
        start_row = title_row_offset + title_height + 1
        header_row = ws[start_row]

        # Locate template "YYYY" column
        template_col_idx = None
        for idx, cell in enumerate(header_row[:10], start=1):
            if cell.value == "YYYY":
                template_col_idx = idx
                break
        if template_col_idx is None:
            raise ValueError("No 'YYYY' template column found in first 10 columns of the KPI sheet.")
        
        # 1) Insert past years until ds_year-1
        ds_year_scholar = ds_year if ds_month >= 9 else ds_year - 1
        month_shifts = [-2,-1,-13]
        insert_idx = template_col_idx + nblank_cols + len(month_shifts) +1

        # Build mappings
        years_mapping, months_mapping = [], []
        mapping_offset = nblank_cols + len(month_shifts)+1
        
        for year in range(min_year, ds_year_scholar):
            ws.insert_cols(insert_idx)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                src_cell = row[template_col_idx - 1]
                dest_cell = row[insert_idx - 1]
                if src_cell.has_style:
                    dest_cell._style = copy(src_cell._style)
                    
            label = f"{year}-{year+1}"
            ws.cell(row=start_row, column=insert_idx, value=label)
            years_mapping.append({int(insert_idx-mapping_offset): label})
            insert_idx += 1

        # 2) Insert blank columns
        for i in range(nblank_cols):
            ws.insert_cols(insert_idx)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                src_cell = row[template_col_idx + i]
                dest_cell = row[insert_idx - 1]
                if src_cell.has_style:
                    dest_cell._style = copy(src_cell._style)
            insert_idx += 1
            
        # 3) Insert month columns from template
        for month in month_shifts:
            target_date = datetime(ds_year, ds_month, 1) + relativedelta(months=month)
            ws.insert_cols(insert_idx)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                src_cell = row[template_col_idx - 1]
                dest_cell = row[insert_idx - 1]
                if src_cell.has_style:
                    dest_cell._style = copy(src_cell._style)
                    
            label = target_date.strftime("%m/%Y")
            ws.cell(row=start_row, column=insert_idx, value=label)
            months_mapping.append({int(insert_idx-mapping_offset): label})
            insert_idx += 1
            
        # 4) Remove template columns
        ws.delete_cols(template_col_idx,nblank_cols + len(month_shifts)+1)
        
        layout["date_mappings"] = {"years": years_mapping, "months": months_mapping}
        self.layout = layout
    
    def _expand_date_columns_kpis(self, ds: str,min_year: int = 2021, nblank_cols: int = 1):
        """Expand date columns in all KPI sheets based on ds (YYYY-MM)."""
        if self.definition =="individual_kpis":
                self._expand_date_columns_indiv(ds,min_year=min_year,nblank_cols=nblank_cols)
        elif self.definition =="collective_kpis":
                self._expand_date_columns_collective(ds,min_year=min_year,nblank_cols=nblank_cols)

    def fill_data(self, duckdb_conn: DuckDBPyConnection, ds: str):
        """Main entry point to fill sheet data based on definition."""
        if self.definition in ("individual_kpis", "collective_kpis"):
            self._fill_kpis(duckdb_conn, ds)
        elif self.definition.startswith("top"):
            pass
            # self._fill_top_sheet(duckdb_conn, ds)

    def _fill_kpis(self, duckdb_conn: DuckDBPyConnection, ds: str):
        """Fill KPI sheets using layout and external query/aggregation helpers."""
        data_source_config = SOURCE_TABLES[SHEET_DEFINITIONS[self.definition].get("source_table")]
        duckdb_table = data_source_config["table"]
        date_mappings = self.layout.get("date_mappings", {})

        # Compute row offset for headers
        min_row = self.layout.get("title_row", 0) + self.layout.get("title_height", 3) + 1

        for row in self.worksheet.iter_rows(min_row=min_row, max_row=self.worksheet.max_row):
            row_config = self._parse_kpi_row(row)
            if not row_config:
                continue

            # Attach date mapping and extra filtering info
            row_config["mappings"] = date_mappings
            row_config["table"] = duckdb_table
            row_config["ds"] = ds

            self._fill_row(duckdb_conn, row_config)

    def _parse_kpi_row(self, row) -> Optional[Dict[str, Any]]:
        """Parse a KPI row in the template."""
        try:
            filter_cell = row[0].value
            if not filter_cell or "=" not in filter_cell:
                return None

            kpi_name = filter_cell.split("=")[1].strip()
            select = row[1].value.strip() if len(row) > 1 and row[1].value else "kpi"
            agg_type = row[2].value if len(row) > 2 else "sum"

            # Determine dimension context
            scale = self.filters.get("scale", "NAT")
            dim_name = "NAT" if scale == "national" else (
                "REG" if scale == "region" else
                "ACAD" if scale == "academy" else
                "DEP" if scale == "department" else "NAT"
            )

            dim_value = self.context.get(scale) if scale != "national" else "NAT"

            return {
                "kpi_name": kpi_name,
                "select_field": select.lower(),
                "agg_type": agg_type,
                "dimension_name": dim_name,
                "dimension_value": dim_value,
            }
        except Exception:
            return None

    def _fill_row(self, duckdb_conn: DuckDBPyConnection, config: Dict[str, Any]):
        """Fill a single row with KPI data using external query & aggregation helpers."""
        mappings = config.get("mappings", {})
        years_mapping = mappings.get("years", [])  # list of single-key dicts
        months_mapping = mappings.get("months", [])
        # extract years
        years_list = [list(d.values())[0] for d in mappings.get("years", [])]
        years_int = [int(y) for y in years_list if str(y).isdigit()]
        years_dict = {list(d.keys())[0]: list(d.values())[0] for d in years_mapping}
        months_dict = {list(d.keys())[0]: list(d.values())[0] for d in months_mapping}

        # Determine start/end years
        years_int = [int(v) for v in years_dict.values() if str(v).isdigit()]
        start_year = min(years_int, default=2021)
        end_year = max(years_int, default=datetime.today().year)
        
        table = config["table"]
        ds = config.get("ds")
        kpi_name = config["kpi_name"]
        dim_name = config["dimension_name"]
        dim_value = config["dimension_value"]
        select_field = config.get("select_field", "kpi")
        agg_type = config.get("agg_type", "sum")
        mappings = config.get("mappings", {})

        scope = "individual" if self.definition == "individual_kpis" else "collective" if self.definition == "collective_kpis" else None
        #typer.echo(f"{kpi_name=},{ds=},{dim_name=},{dim_value=},{select_field=},{agg_type=},{start_year=},{end_year=},{scope=}")
        # 1) Query yearly data
        yearly_data = query_yearly_kpi(
            conn=duckdb_conn,
            ds=ds,
            kpi_name=kpi_name,
            table_name=table,
            dimension_name=dim_name,
            dimension_value=dim_value,
            start_year=start_year,
            end_year=end_year,
            scope=scope,
        )
        typer.echo(f"{agg_type=},{yearly_data=},{select_field=},{scope=}")
        yearly_agg = aggregate_kpi_data(yearly_data, agg_type, "yearly", select_field,scope)

        # 2) Query monthly data
        monthly_data = query_monthly_kpi(
            conn=duckdb_conn,
            table_name=table,
            kpi_name=kpi_name,
            dimension_name=dim_name,
            dimension_value=dim_value,
            ds=ds,
        )

        # monthly_agg = aggregate_kpi_data(monthly_data, agg_type, "monthly", select_field)

        # 3) Write yearly data
        for col_idx, year_label in years_dict.items():
            try:
                value = yearly_agg.get(int(year_label), None)
            except ValueError:
                value = None
            if value is not None:
                typer.echo(f"Writing {value} at row {config.get('row_idx', 0)+1}, col {col_idx+1} for year {year_label}")
                self.worksheet.cell(row=config.get("row_idx", 0) + 1, column=col_idx + 1, value=value)

        # 4) Write monthly data
        for col_idx, month_label in months_dict.items():
            if month_label in monthly_data:
                self.worksheet.cell(row=config.get("row_idx", 0) + 1, column=col_idx + 1, value=monthly_data[month_label])
                
    def _fill_top_sheet(self, duckdb_conn: DuckDBPyConnection, ds: str):
        """Fill top sheets using same logic but adapted to top-specific filters."""
        # similar logic as _fill_kpis, but use self.tab_name/scale for dimension
        pass
    # def fill_data(self,duckdb_conn: DuckDBPyConnection):
    #     """Placeholder for filling the sheet with data based on filters."""
    #     if self.definition == "individual_kpis":
    #         self._fill_individual_kpis(duckdb_conn)
    #     elif self.definition == "collective_kpis":
    #         self._fill_collective_kpis(duckdb_conn)
    #     elif self.definition.startswith("top"):
    #         self._fill_top_sheet(duckdb_conn)
    
    # def _fill_kpis(self,duckdb_conn: DuckDBPyConnection):
    #     """Fill individual_kpis sheet with data."""
    #     data_source_config = SOURCE_TABLES[SHEET_DEFINITIONS[self.definition].get("source_table")]
    #     duckdb_table = data_source_config["table"]
    #     date_mappings = self.layout.get("date_mappings", {})

    #     min_row = self.layout.get("title_row_offset", 0) + self.layout.get("title_height", 3) + 1
    #     for row in self.worksheet.iter_rows(min_row=min_row, max_row=self.worksheet.max_row):
    #         query_config = self._parse_row(row)
    #         query_config["mappings"] = date_mappings
    #         extra_config = date_mappings | ## TIME FILTERING STUFF
    #         if query_config:
    #             self._fill_row(query_config | extra_config)
            
    #     #iterate on rows to build queries and fill data !
    #     pass
    
    # def _parse_row(self,row):
    #     kpi_name = row[0].value.split("=")[1].strip()
    #     if not kpi_name or kpi_name == "":
    #         return None
    #     select = row[1].value.strip()
    #     agg_type = row[2].value
    #     dim_name = self.filters.get("scale")
    #     dim_value = self.context.get(self.context.get("scale")) if dim_name != 'national' else 'NAT'
    #     return {
    #         "kpi_name": kpi_name,
    #         "select": select,
    #         "agg_type": agg_type,
    #         "dimension_name": 'NAT' if dim_name == 'national' else 'REG' if dim_name == 'region' else 'ACAD' if dim_name == 'academy' else 'DEP' if dim_name == 'department' else None,
    #         "dimension_value": dim_value,
    #     }
    # def _fill_row(self,row,duckdb_conn,query_config):
    #     pass
    # def build_query(self,config):
    #     base_query = f"""
    #     SELECT {config['select']} as kpi
    #     FROM {config['table']} 
    #     WHERE kpi_name = '{config['kpi_name']}'
    #     """
    #     if config['dimension_name'] and config['dimension_value']:
    #         base_query += f" AND dimension_name = '{config['dimension_name']}' AND dimension_value = '{config['dimension_value']}'"
    #     return base_query
    
    # def _parse_template_for_kpis(self) -> List[Dict[str, Any]]:
    #     """Parse the worksheet to identify KPI rows and their configurations."""
    #     base_row_idx = self.layout.get("title_row_offset", 0) + self.layout.get("title_height", 3) + 1
    #     kpi_rows = []
    #     ws = self.worksheet
    #     layout = SHEET_LAYOUT.get("kpis", {})
    #     title_row_offset = layout.get("title_row", 0)
    #     title_height = layout.get("title_height", 3)
    #     start_row = title_row_offset + title_height + 1
        
    #     for row in ws.iter_rows(min_row=start_row + 1, max_row=ws.max_row):
    #         kpi_name = row[0].value
    #         if not kpi_name:
    #             continue
    
    
    def _fill_collective_kpis(self,duckdb_conn: DuckDBPyConnection):
        """Fill collective_kpis sheet with data."""
        pass    
    
    def _fill_top_sheet(self,duckdb_conn: DuckDBPyConnection):  
        """Fill top sheet with data."""
        pass
    def postprocess(self):
        """Placeholder for postprocessing formatting."""
        pass



class Report:
    """Represents a single Excel report for a stakeholder."""
    
    def __init__(self, report_type: str, stakeholder, base_template_path: Path, output_path: Path, context: Dict[str, Any] = None):
        self.report_type = report_type
        self.stakeholder = stakeholder
        self.base_template_path = base_template_path
        self.output_path = output_path
        self.context = context or {}
        self.workbook: Any = None          # openpyxl.Workbook
        self.sheets: List[Sheet] = []      # list of Sheet instances

    def _load_template(self):
        """Load the template workbook directly as the report workbook."""
        self.workbook = openpyxl.load_workbook(self.base_template_path)
        
        # Remove default sheet if it exists and is empty
        if "Sheet" in self.workbook.sheetnames and len(self.workbook["Sheet"].rows) == 0:
            self.workbook.remove(self.workbook["Sheet"])

    def _copy_template_tab(self, template_tab: str, new_tab_name: str):
        template_ws = self.workbook_template[template_tab]
        new_ws = self.workbook.copy_worksheet(template_ws)
        new_ws.title = new_tab_name[:31]  # truncate if needed
        return new_ws
    
    def _build_tab_name(self, definition_key: str, filters: Dict[str, Any], context: Dict[str, Any]):
        """Return a user-friendly tab name based on sheet definition, filters, and context."""
        
        if definition_key == "individual_kpis":
            base = "Part Individuelle"
            scale = filters.get("scale", "").capitalize()
            return f"{base} ({scale})"
        
        elif definition_key == "collective_kpis":
            base = "Part Collective"
            scale = filters.get("scale", "").capitalize()
            return f"{base} ({scale})"
    
        else:
            tab_name = SHEET_DEFINITIONS.get(definition_key, {}).get("tab_name")
            if tab_name is None:
                tab_name = SHEET_DEFINITIONS.get(definition_key, {}).get("template_tab", definition_key)
            return tab_name.capitalize()
    
    def _resolve_sheets(self):
        """Rename and configure tabs from the template instead of copying across workbooks."""
        blueprint = REPORTS[self.report_type]
        for sheet_info in blueprint["sheets"]:
            definition_key = sheet_info["definition"]
            template_info = SHEET_DEFINITIONS[definition_key]
            
            # Decide tab name based on definition + filters + context
            tab_name = self._build_tab_name(
                definition_key, 
                {**sheet_info.get("filters", {}), **self.context},
                self.context
            )

            # Check if the tab already exists
            if tab_name in self.workbook.sheetnames:
                ws = self.workbook[tab_name]
            else:
                # Copy the template tab
                template_ws = self.workbook[template_info["template_tab"]]
                ws = self.workbook.copy_worksheet(template_ws)
       
            if len(tab_name) > 31:
                typer.secho(f"⚠️ Tab name '{tab_name}' exceeds 31 characters, truncating.", fg="yellow")
                tab_name = tab_name[:31] # truncate if too long
            ws.title = tab_name  
            
            # Create Sheet instance
            sheet = Sheet(
                definition=definition_key,
                template_tab=template_info["template_tab"],
                worksheet=ws,
                filters={**sheet_info.get("filters", {}), **self.context},
                tab_name=tab_name,
                stakeholder=self.stakeholder,
                context=self.context
            )
            self.sheets.append(sheet)

    def _cleanup_template_sheets(self, used_tabs: List[str]):
        """
        Remove all template/unused sheets and enforce blueprint order.
        """
        # Remove any sheet not in the used list
        for sheet_name in list(self.workbook.sheetnames):
            if sheet_name not in used_tabs:
                self.workbook.remove(self.workbook[sheet_name])

        # Enforce order from used_tabs
        for idx, tab_name in enumerate(used_tabs):
            ws = self.workbook[tab_name]
            self.workbook._sheets.remove(ws)
            self.workbook._sheets.insert(idx, ws)
    
    def build(self,ds: str, duckdb_conn: DuckDBPyConnection):
        """Build the report workbook with all sheets (copy templates, preprocess, fill data)."""
        self._load_template()
        self._resolve_sheets()
        self._cleanup_template_sheets([sheet.tab_name for sheet in self.sheets])

        # Preprocess and fill data for each sheet
        for sheet in self.sheets:
            sheet.preprocess(ds)
            sheet.fill_data(duckdb_conn,ds)

    def save(self):
        """Save the workbook to the output path."""
        if not self.output_path.suffix == ".xlsx":
            self.output_path = self.output_path.with_suffix(".xlsx")
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(self.output_path)
        
class ReportPlanner:
    """Expands a stakeholder into the list of reports they should get."""

    def __init__(self, stakeholder):
        self.stakeholder = stakeholder

    def plan_reports(self) -> List[Dict[str, Any]]:
        """
        Returns a list of report jobs:
        {
            "report_type": str,
            "context": dict,  # region, academy, department, etc.
            "output_path": Path
        }
        """
        jobs = []

        if self.stakeholder.type.name == "MINISTERE":
            # Ministère gets only the national summary
            # base_path = Path("NATIONAL")
            jobs.append(
                {
                    "report_type": "national_summary",
                    "context": {},
                    "output_path": "rapport_national.xlsx"
                }
            )
            return jobs

        # DRAC gets:
        # 1) Regional summary (includes national for comparison)
        region_name = self.stakeholder.name
        # base_path = Path("REGIONAL") / f"{region_name}"
        jobs.append(
            {
                "report_type": "region_summary",
                "context": {"region": region_name},
                "output_path": "rapport_regional.xlsx"
            }
            )

        # 2) One detailed report per académie (collective scope)
        if self.stakeholder.academy_tree:
            for node in self.stakeholder.academy_tree.all_nodes():
                if node.tag not in ["root_acad", "Academies"]:
                    jobs.append(
                        {
                            "report_type": "academy_detail",
                            "context": {"region": region_name, "academy": node.tag},
                            "output_path": f"academie_{node.tag}.xlsx"
                        }
                    )

        # 3) One detailed report per département (individual scope)
        if self.stakeholder.department_tree:
            for node in self.stakeholder.department_tree.all_nodes():
                if node.tag not in ["root_dept", "Departements"]:
                    jobs.append(
                        {
                            "report_type": "department_detail",
                            "context": {"region": region_name, "department": node.tag},
                            "output_path": f"departement_{node.tag}.xlsx"
                        }
                    )

        return jobs